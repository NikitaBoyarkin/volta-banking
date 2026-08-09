"""
Report Generator Module

Utilities for generating Excel, PDF, and Word reports from analysis data.
Based on Anthropic Skills patterns for document automation.

Usage:
    from utils.report_generator import generate_excel_report, generate_pdf_summary
"""

import os
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# =============================================================================
# Color Constants (Financial Industry Standard)
# =============================================================================
COLOR_BLUE = "0000FF"  # Hardcoded inputs
COLOR_BLACK = "000000"  # Formulas
COLOR_GREEN = "008000"  # Cross-sheet links
COLOR_RED = "FF0000"  # External links
COLOR_YELLOW = "FFFF00"  # Key assumptions/highlights
COLOR_GRAY = "F0F0F0"  # Alternating rows


def generate_excel_report(
    data_dict: dict,
    output_path: str,
    title: str = "Analysis Report",
    author: str = "Data Analytics Team",
) -> str:
    """
    Generate Excel report with multiple sheets and professional formatting.

    Parameters
    ----------
    data_dict : dict
        Dictionary of {sheet_name: DataFrame}
    output_path : str
        Output file path
    title : str
        Report title for summary sheet
    author : str
        Author name for metadata

    Returns
    -------
    str
        Path to generated file

    Examples
    --------
    >>> data = {
    ...     'Summary': pd.DataFrame({'Metric': ['Users', 'Revenue'], 'Value': [1000, 50000]}),
    ...     'Details': df.groupby('category').sum()
    ... }
    >>> generate_excel_report(data, 'report.xlsx', title='Q4 Analysis')
    'report.xlsx'
    """

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Add summary sheet first
    summary_data = _create_summary_sheet(data_dict, title, author)
    _add_sheet(wb, summary_data, "Summary", is_summary=True)

    # Add data sheets
    for sheet_name, df in data_dict.items():
        # Clean sheet name (Excel has 31 char limit)
        clean_name = str(sheet_name)[:31]
        _add_sheet(wb, df, clean_name)

    # Save
    os.makedirs(
        os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True
    )
    wb.save(output_path)

    print(f"✓ Report saved to {output_path}")
    return output_path


def _create_summary_sheet(data_dict: dict, title: str, author: str) -> pd.DataFrame:
    """Create summary sheet with metadata and overview."""

    summary_rows = [
        ["Report Title", title],
        ["Generated", datetime.now().strftime("%Y-%m-%d %H:%M")],
        ["Author", author],
        ["", ""],
        ["Sheets Included", ""],
    ]

    for sheet_name, df in data_dict.items():
        summary_rows.append([f"  • {sheet_name}", f"{len(df)} rows"])

    summary_rows.extend(
        [
            ["", ""],
            ["Quick Stats", ""],
        ]
    )

    # Add row counts
    total_rows = sum(len(df) for df in data_dict.values())
    summary_rows.append(["Total Data Rows", total_rows])
    summary_rows.append(["Number of Sheets", len(data_dict)])

    return pd.DataFrame(summary_rows, columns=["Item", "Value"])


def _add_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str, is_summary: bool = False) -> None:
    """Add a formatted sheet to workbook."""

    ws = wb.create_sheet(title=sheet_name)

    # Add data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Header formatting
            if r_idx == 1:
                cell.font = Font(bold=True, color=COLOR_BLUE)
                cell.fill = PatternFill(start_color=COLOR_YELLOW, fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Alternating row colors (not for summary)
            elif not is_summary and r_idx % 2 == 0:
                cell.fill = PatternFill(start_color=COLOR_GRAY, fill_type="solid")

            # Number formatting
            if isinstance(value, (int, float)) and r_idx > 1:
                if abs(value) >= 1000000:
                    cell.number_format = '$#,##0,,"M"'
                elif abs(value) >= 1000:
                    cell.number_format = "$#,##0"
                elif 0 < abs(value) < 1:
                    cell.number_format = "0.00%"
                else:
                    cell.number_format = "#,##0"

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_funnel_excel(
    funnel_data: pd.DataFrame,
    segment_data: pd.DataFrame,
    output_path: str = "output/funnel_analysis.xlsx",
) -> str:
    """
    Generate Excel report specifically for funnel analysis.

    Parameters
    ----------
    funnel_data : pd.DataFrame
        Funnel step data with columns: stage, users, conversion_rate
    segment_data : pd.DataFrame
        Segment breakdown data
    output_path : str
        Output file path

    Returns
    -------
    str
        Path to generated file

    Examples
    --------
    >>> funnel_df = pd.DataFrame({
    ...     'stage': ['Visited', 'Signed Up', 'Activated'],
    ...     'users': [10000, 5000, 2000],
    ...     'conversion_rate': [100, 50, 20]
    ... })
    >>> generate_funnel_excel(funnel_df, segment_df, 'funnel_report.xlsx')
    """

    data_dict = {"Funnel": funnel_data, "Segments": segment_data}

    return generate_excel_report(
        data_dict, output_path, title="Funnel Analysis Report", author="Product Analytics Team"
    )


def generate_rfm_excel(
    rfm_data: pd.DataFrame,
    segment_summary: pd.DataFrame,
    output_path: str = "output/rfm_analysis.xlsx",
) -> str:
    """
    Generate Excel report for RFM segmentation analysis.

    Parameters
    ----------
    rfm_data : pd.DataFrame
        Full RFM data with customer segments
    segment_summary : pd.DataFrame
        Segment-level summary statistics
    output_path : str
        Output file path

    Returns
    -------
    str
        Path to generated file
    """

    # Create summary statistics
    summary_df = pd.DataFrame(
        {
            "Metric": ["Total Customers", "Segments", "Avg RFM Score"],
            "Value": [
                len(rfm_data),
                rfm_data["Segment"].nunique(),
                round(rfm_data["RFMScore"].mean(), 2),
            ],
        }
    )

    data_dict = {
        "Summary": summary_df,
        "Segments": segment_summary,
        "Customer_RFM": rfm_data[
            ["customer_id", "R_Quartile", "F_Quartile", "M_Quartile", "Segment"]
        ],
    }

    return generate_excel_report(
        data_dict, output_path, title="RFM Segmentation Analysis", author="Customer Analytics Team"
    )


def generate_ab_test_excel(
    test_results: dict, output_path: str = "output/ab_test_results.xlsx"
) -> str:
    """
    Generate Excel report for A/B test results.

    Parameters
    ----------
    test_results : dict
        Dictionary with test results:
        - summary: DataFrame with metric comparison
        - detailed: DataFrame with daily breakdown
        - segments: DataFrame with segment analysis
    output_path : str
        Output file path

    Returns
    -------
    str
        Path to generated file
    """

    data_dict = {
        "Summary": test_results.get("summary", pd.DataFrame()),
        "Daily": test_results.get("daily", pd.DataFrame()),
        "Segments": test_results.get("segments", pd.DataFrame()),
    }

    return generate_excel_report(
        data_dict, output_path, title="A/B Test Results", author="Experimentation Team"
    )


# =============================================================================
# PDF Generation (Placeholder for future implementation)
# =============================================================================


def generate_pdf_summary(
    analysis_results: dict, output_path: str = "output/analysis_summary.pdf"
) -> str:
    """
    Generate PDF summary report.

    Note: Requires additional dependencies (reportlab or weasyprint).
    This is a placeholder for future implementation.

    Parameters
    ----------
    analysis_results : dict
        Dictionary with analysis results
    output_path : str
        Output file path

    Returns
    -------
    str
        Path to generated file
    """

    # TODO: Implement PDF generation using reportlab
    # For now, return message
    print("⚠ PDF generation not yet implemented.")
    print("  Install reportlab: pip install reportlab")
    print("  Or use Excel report: generate_excel_report()")

    return output_path


# =============================================================================
# Word Document Generation (Placeholder for future implementation)
# =============================================================================


def generate_word_report(
    analysis_results: dict,
    template_path: str = None,
    output_path: str = "output/analysis_report.docx",
) -> str:
    """
    Generate Word document report.

    Note: Requires python-docx library.
    This is a placeholder for future implementation.

    Parameters
    ----------
    analysis_results : dict
        Dictionary with analysis results
    template_path : str
        Path to Word template (optional)
    output_path : str
        Output file path

    Returns
    -------
    str
        Path to generated file
    """

    # TODO: Implement Word generation using python-docx
    print("⚠ Word generation not yet implemented.")
    print("  Install python-docx: pip install python-docx")
    print("  Or use Excel report: generate_excel_report()")

    return output_path


if __name__ == "__main__":
    # Example usage
    sample_data = {
        "Metrics": pd.DataFrame(
            {"Metric": ["Users", "Revenue", "Conversion"], "Value": [10000, 500000, 0.15]}
        ),
        "Details": pd.DataFrame(
            {
                "Category": ["A", "B", "C"],
                "Count": [100, 200, 150],
                "Revenue": [10000, 20000, 15000],
            }
        ),
    }

    generate_excel_report(sample_data, "output/sample_report.xlsx")
